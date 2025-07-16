import serial
import pandas as pd
from datetime import datetime
import time
import os
from openpyxl import load_workbook, Workbook

# Threshold constants (matching Arduino thresholds)
VOLTAGE_THRESHOLD = 5.5  # Voltage threshold in volts
TEMP_THRESHOLD = 30.0   # Temperature threshold in Celsius

# Excel file names
excel_file = "sensor_data.xlsx"
MASTER_EXCEL_FILE = "master_data.xlsx"  # Added for Master
BACKUP_EXCEL_FILE = "backup_data.xlsx"  # Added for Backup

# Expected columns for the Excel file
expected_columns = ["Timestamp", "Device", "Voltage (V)", "Temperature (C)", "Status"]

# Initialize Excel file with two sheets if it doesn't exist or is incompatible
def initialize_excel_file(file_path):
    df_normal = pd.DataFrame(columns=expected_columns)
    df_error = pd.DataFrame(columns=expected_columns)
    try:
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df_normal.to_excel(writer, sheet_name='Sheet1', index=False)
            df_error.to_excel(writer, sheet_name='ErrorSheet', index=False)
        print(f"Created new Excel file: {file_path}")
    except PermissionError:
        print(f"Permission denied: Cannot write to '{file_path}'. Ensure the file is not open in another program.")
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        new_file = f"sensor_data_{timestamp}.xlsx"
        print(f"Creating fallback file: {new_file}")
        with pd.ExcelWriter(new_file, engine='openpyxl') as writer:
            df_normal.to_excel(writer, sheet_name='Sheet1', index=False)
            df_error.to_excel(writer, sheet_name='ErrorSheet', index=False)
        return new_file
    return file_path

# Initialize Excel file with two sheets (Readings and Errors) for Master and Backup
def initialize_specific_excel_file(file_path):
    df_readings = pd.DataFrame(columns=expected_columns)
    df_error = pd.DataFrame(columns=expected_columns + ["Error Link"])
    try:
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df_readings.to_excel(writer, sheet_name='Readings', index=False)
            df_error.to_excel(writer, sheet_name='Errors', index=False)
        print(f"Created new Excel file: {file_path}")
    except PermissionError:
        print(f"Permission denied: Cannot write to '{file_path}'. Ensure the file is not open in another program.")
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        new_file = f"{file_path.split('.')[0]}_{timestamp}.xlsx"
        print(f"Creating fallback file: {new_file}")
        with pd.ExcelWriter(new_file, engine='openpyxl') as writer:
            df_readings.to_excel(writer, sheet_name='Readings', index=False)
            df_error.to_excel(writer, sheet_name='Errors', index=False)
        return new_file
    return file_path

# Check and load Excel file (original, kept for structure compatibility)
if not os.path.exists(excel_file):
    excel_file = initialize_excel_file(excel_file)
    df_normal = pd.DataFrame(columns=expected_columns)
    df_error = pd.DataFrame(columns=expected_columns)
else:
    try:
        workbook = load_workbook(excel_file)
        sheet_names = workbook.sheetnames
        if 'ErrorSheet' not in sheet_names:
            print("Warning: 'ErrorSheet' not found. Adding it to the Excel file.")
            df_error = pd.DataFrame(columns=expected_columns)
            with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
                df_error.to_excel(writer, sheet_name='ErrorSheet', index=False)
        else:
            df_error = pd.read_excel(excel_file, sheet_name='ErrorSheet')

        df_normal = pd.read_excel(excel_file, sheet_name='Sheet1')
        if list(df_normal.columns) != expected_columns or list(df_error.columns) != expected_columns:
            print("Warning: Excel file has incompatible columns. Creating a new file.")
            excel_file = initialize_excel_file(excel_file)
            df_normal = pd.DataFrame(columns=expected_columns)
            df_error = pd.DataFrame(columns=expected_columns)
    except Exception as e:
        print(f"Error reading Excel file: {e}. Creating a new file.")
        excel_file = initialize_excel_file(excel_file)
        df_normal = pd.DataFrame(columns=expected_columns)
        df_error = pd.DataFrame(columns=expected_columns)

# Check and load Excel file for Master
if not os.path.exists(MASTER_EXCEL_FILE):
    master_excel_file = initialize_specific_excel_file(MASTER_EXCEL_FILE)
    df_master_readings = pd.DataFrame(columns=expected_columns)
    df_master_error = pd.DataFrame(columns=expected_columns + ["Error Link"])
else:
    try:
        workbook = load_workbook(MASTER_EXCEL_FILE)
        sheet_names = workbook.sheetnames
        if 'Errors' not in sheet_names:
            print("Warning: 'Errors' not found in Master file. Adding it.")
            df_master_error = pd.DataFrame(columns=expected_columns + ["Error Link"])
            with pd.ExcelWriter(MASTER_EXCEL_FILE, engine='openpyxl', mode='a') as writer:
                df_master_error.to_excel(writer, sheet_name='Errors', index=False)
        else:
            df_master_error = pd.read_excel(MASTER_EXCEL_FILE, sheet_name='Errors')

        df_master_readings = pd.read_excel(MASTER_EXCEL_FILE, sheet_name='Readings')
        if list(df_master_readings.columns) != expected_columns or list(df_master_error.columns) != (expected_columns + ["Error Link"]):
            print("Warning: Master Excel file has incompatible columns. Creating a new file.")
            master_excel_file = initialize_specific_excel_file(MASTER_EXCEL_FILE)
            df_master_readings = pd.DataFrame(columns=expected_columns)
            df_master_error = pd.DataFrame(columns=expected_columns + ["Error Link"])
    except Exception as e:
        print(f"Error reading Master Excel file: {e}. Creating a new file.")
        master_excel_file = initialize_specific_excel_file(MASTER_EXCEL_FILE)
        df_master_readings = pd.DataFrame(columns=expected_columns)
        df_master_error = pd.DataFrame(columns=expected_columns + ["Error Link"])

# Check and load Excel file for Backup
if not os.path.exists(BACKUP_EXCEL_FILE):
    backup_excel_file = initialize_specific_excel_file(BACKUP_EXCEL_FILE)
    df_backup_readings = pd.DataFrame(columns=expected_columns)
    df_backup_error = pd.DataFrame(columns=expected_columns + ["Error Link"])
else:
    try:
        workbook = load_workbook(BACKUP_EXCEL_FILE)
        sheet_names = workbook.sheetnames
        if 'Errors' not in sheet_names:
            print("Warning: 'Errors' not found in Backup file. Adding it.")
            df_backup_error = pd.DataFrame(columns=expected_columns + ["Error Link"])
            with pd.ExcelWriter(BACKUP_EXCEL_FILE, engine='openpyxl', mode='a') as writer:
                df_backup_error.to_excel(writer, sheet_name='Errors', index=False)
        else:
            df_backup_error = pd.read_excel(BACKUP_EXCEL_FILE, sheet_name='Errors')

        df_backup_readings = pd.read_excel(BACKUP_EXCEL_FILE, sheet_name='Readings')
        if list(df_backup_readings.columns) != expected_columns or list(df_backup_error.columns) != (expected_columns + ["Error Link"]):
            print("Warning: Backup Excel file has incompatible columns. Creating a new file.")
            backup_excel_file = initialize_specific_excel_file(BACKUP_EXCEL_FILE)
            df_backup_readings = pd.DataFrame(columns=expected_columns)
            df_backup_error = pd.DataFrame(columns=expected_columns + ["Error Link"])
    except Exception as e:
        print(f"Error reading Backup Excel file: {e}. Creating a new file.")
        backup_excel_file = initialize_specific_excel_file(BACKUP_EXCEL_FILE)
        df_backup_readings = pd.DataFrame(columns=expected_columns)
        df_backup_error = pd.DataFrame(columns=expected_columns + ["Error Link"])

# Serial port setup
SERIAL_PORT_MASTER = 'COM5'  # Master port
SERIAL_PORT_BACKUP = 'COM2'  # Backup port
BAUD_RATE = 9600
TIMEOUT_MASTER = 5  # 5 seconds to determine if Master has failed

# Variables to track Master state
last_data_time = datetime.now()
master_alive = True
current_serial = None  # To hold the active serial connection

def connect_serial(port):
    """Attempt to connect to a serial port."""
    try:
        ser = serial.Serial(port, BAUD_RATE, timeout=1)
        print(f"Connected to {port} at {BAUD_RATE} baud")
        time.sleep(5)  # Wait for Arduino to stabilize
        ser.flushInput()  # Clear input buffer
        ser.flushOutput()  # Clear output buffer
        return ser
    except serial.SerialException as e:
        print(f"Error opening serial port {port}: {e}")
        return None

# Initialize serial connection to Master
current_serial = connect_serial(SERIAL_PORT_MASTER)
if current_serial is None:
    current_serial = connect_serial(SERIAL_PORT_BACKUP)
    if current_serial is None:
        print("Failed to connect to Backup. Exiting...")
        exit()
    else:
        master_alive = False
else:
    master_alive = True

print(f"Listening to {'Master' if master_alive else 'Backup'}... Press Ctrl+C to stop.")

def log_raw_data(timestamp, line, is_bytes=False):
    """Log raw serial data to file, handling non-text bytes safely."""
    try:
        with open('raw_serial_log.txt', 'a', encoding='utf-8') as f:
            if is_bytes:
                hex_line = line.hex()
                f.write(f"{timestamp} | Raw Bytes: {hex_line}\n")
            else:
                f.write(f"{timestamp} | Raw: {line}\n")
    except UnicodeEncodeError as e:
        print(f"Warning: Could not log raw data: {e}")
        with open('raw_serial_log.txt', 'a', encoding='utf-8') as f:
            f.write(f"{timestamp} | Raw: [Unprintable data]\n")
    except Exception as e:
        print(f"Error logging raw data: {e}")

try:
    # Clear initial garbage data
    for _ in range(15):  # Read up to 15 lines to skip startup noise
        if current_serial.in_waiting > 0:
            current_serial.readline()
        time.sleep(0.1)

    while True:
        if current_serial.in_waiting > 0:
            try:
                # Read raw bytes
                raw_line = current_serial.readline()
                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                # Update last data time if we received data
                last_data_time = datetime.now()

                # Log raw bytes as hex
                log_raw_data(timestamp, raw_line, is_bytes=True)

                # Try decoding as utf-8 first, fallback to latin1
                try:
                    line = raw_line.decode('utf-8', errors='ignore').strip()
                except UnicodeDecodeError:
                    line = raw_line.decode('latin1', errors='ignore').strip()

                if line:
                    # Log decoded line
                    log_raw_data(timestamp, line)

                    # Process only sensor data lines
                    if "Data: Backup" in line or "Data: Slave" in line or "Data: Master" in line:
                        # Print the sensor data line
                        print(f"{line}")

                        # Parse the data
                        try:
                            parts = line.split(',')
                            device = "Backup" if "Data: Backup" in line else ("Master" if "Data: Master" in line else "Slave")
                            voltage_str = parts[1].split(":")[1].strip().split()[0]  # Extract voltage number
                            temp_str = parts[2].split(":")[1].strip().split()[0]    # Extract temperature number

                            voltage_val = float(voltage_str)
                            temp_val = float(temp_str)

                            # Determine status
                            status = "Normal"
                            if voltage_val > VOLTAGE_THRESHOLD:
                                status = "Overvoltage"
                            elif temp_val > TEMP_THRESHOLD:
                                status = "Overheat"

                            new_row = {
                                "Timestamp": timestamp,
                                "Device": device,
                                "Voltage (V)": voltage_val,
                                "Temperature (C)": temp_val,
                                "Status": status
                            }

                            # Determine which Excel file to use based on the source
                            if master_alive:
                                # Add to Master Readings (all data, including Backup and Slave while Master is alive)
                                if not df_master_readings.empty or not new_row.empty:
                                    df_master_readings = pd.concat([df_master_readings, pd.DataFrame([new_row])], ignore_index=True)

                                # Add to Master Errors if status is not Normal
                                if status != "Normal":
                                    row_idx = len(df_master_readings)  # Row index in Readings (1-based, including header)
                                    new_row["Error Link"] = f"=HYPERLINK(\"#Readings!A{row_idx}\", \"Go to Reading\")"
                                    if not df_master_error.empty or not new_row.empty:
                                        df_master_error = pd.concat([df_master_error, pd.DataFrame([new_row])], ignore_index=True)

                                # Save to Master Excel
                                try:
                                    with pd.ExcelWriter(MASTER_EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                                        df_master_readings.to_excel(writer, sheet_name='Readings', index=False)
                                        df_master_error.to_excel(writer, sheet_name='Errors', index=False)
                                    print(f"Saved Master data: {new_row} to {'Readings' if status == 'Normal' else 'Readings and Errors'}")
                                except PermissionError:
                                    print(f"Permission denied: Cannot write to '{MASTER_EXCEL_FILE}'. Ensure the file is not open.")
                                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                                    new_file = f"master_data_{timestamp}.xlsx"
                                    print(f"Saving to fallback file: {new_file}")
                                    with pd.ExcelWriter(new_file, engine='openpyxl') as writer:
                                        df_master_readings.to_excel(writer, sheet_name='Readings', index=False)
                                        df_master_error.to_excel(writer, sheet_name='Errors', index=False)
                                    master_excel_file = new_file
                                except Exception as e:
                                    print(f"Error saving to Master Excel: {e}")
                                    log_raw_data(timestamp, f"Excel Save Error: {e}")
                            else:
                                # Add to Backup Readings (for both Backup and Slave after Master fails)
                                if not df_backup_readings.empty or not new_row.empty:
                                    df_backup_readings = pd.concat([df_backup_readings, pd.DataFrame([new_row])], ignore_index=True)

                                # Add to Backup Errors if status is not Normal
                                if status != "Normal":
                                    row_idx = len(df_backup_readings)  # Row index in Readings (1-based, including header)
                                    new_row["Error Link"] = f"=HYPERLINK(\"#Readings!A{row_idx}\", \"Go to Reading\")"
                                    if not df_backup_error.empty or not new_row.empty:
                                        df_backup_error = pd.concat([df_backup_error, pd.DataFrame([new_row])], ignore_index=True)

                                # Save to Backup Excel
                                try:
                                    with pd.ExcelWriter(BACKUP_EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                                        df_backup_readings.to_excel(writer, sheet_name='Readings', index=False)
                                        df_backup_error.to_excel(writer, sheet_name='Errors', index=False)
                                    print(f"Saved Backup/Slave data: {new_row} to {'Readings' if status == 'Normal' else 'Readings and Errors'}")
                                except PermissionError:
                                    print(f"Permission denied: Cannot write to '{BACKUP_EXCEL_FILE}'. Ensure the file is not open.")
                                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                                    new_file = f"backup_data_{timestamp}.xlsx"
                                    print(f"Saving to fallback file: {new_file}")
                                    with pd.ExcelWriter(new_file, engine='openpyxl') as writer:
                                        df_backup_readings.to_excel(writer, sheet_name='Readings', index=False)
                                        df_backup_error.to_excel(writer, sheet_name='Errors', index=False)
                                    backup_excel_file = new_file
                                except Exception as e:
                                    print(f"Error saving to Backup Excel: {e}")
                                    log_raw_data(timestamp, f"Excel Save Error: {e}")

                        except (IndexError, ValueError) as e:
                            print(f"Error parsing line '{line}': {e}")
                            log_raw_data(timestamp, f"Parse Error: {e}")

            except serial.SerialException as e:
                print(f"Serial error: {e}")
                log_raw_data(datetime.now().strftime('%Y-%m-%d %H:%M:%S'), f"Serial Error: {e}")
                break
            except Exception as e:
                print(f"Unexpected error: {e}")
                log_raw_data(datetime.now().strftime('%Y-%m-%d %H:%M:%S'), f"Unexpected Error: {e}")

        # Check if Master has failed (no data for TIMEOUT_MASTER seconds)
        if master_alive and (datetime.now() - last_data_time).total_seconds() > TIMEOUT_MASTER:
            current_serial.close()
            current_serial = connect_serial(SERIAL_PORT_BACKUP)
            if current_serial is None:
                print("Failed to connect to Backup. Exiting...")
                break
            else:
                master_alive = False
                print("Switched to Backup. Listening to Backup...")
                # Clear initial garbage data from Backup
                for _ in range(15):
                    if current_serial.in_waiting > 0:
                        current_serial.readline()
                    time.sleep(0.1)

        time.sleep(0.1)  # Reduced delay for responsiveness

except KeyboardInterrupt:
    print("Stopped listening.")
finally:
    if current_serial:
        current_serial.close()
    print("Serial port closed")